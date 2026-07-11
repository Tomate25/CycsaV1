import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\abdia\Downloads\Exportar_data_clientes_29-06-2026 (1).xlsx", read_only=True)
sheet = wb.active

# Find header row
header_row_idx = None
headers = []

for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
    # Look for a row that has "Tipo de Cliente" or "Id (No Modificar)"
    row_str = [str(x).strip() if x is not None else "" for x in row]
    if "Tipo de Cliente" in row_str or "Id (No Modificar)" in row_str or "Nombre del Cliente" in row_str:
        header_row_idx = r_idx
        headers = row_str
        break

if header_row_idx is not None:
    print(f"Header row found at index {header_row_idx}:")
    for col_idx, h in enumerate(headers):
        print(f"Col {col_idx}: '{h}'")
        
    # Read first 3 data rows
    print("\nFIRST 3 DATA ROWS:")
    data_count = 0
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx <= header_row_idx:
            continue
        print(f"Row {r_idx}: {row[:10]} ...")
        data_count += 1
        if data_count >= 3:
            break
else:
    print("Header row NOT found. Let's dump first 15 rows completely:")
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx >= 15:
            break
        print(f"Row {r_idx}: {row[:5]}")
